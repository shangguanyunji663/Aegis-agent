from __future__ import annotations

import json
import smtplib
import ssl
import threading
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Any
from urllib import request

from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.privacy import redacted_payload
from app.tool_contracts import normalize_tool_kind


EXCEL_WRITE_LOCK = threading.Lock()


class ToolExecutionService:
    def __init__(self, settings: Settings):
        self.settings = settings

    def execute(self, kind: str, payload: dict[str, Any], attempts: int) -> dict[str, Any]:
        canonical_kind = normalize_tool_kind(kind)
        if payload.get("always_fail"):
            raise RuntimeError(f"{canonical_kind} forced failure")
        if payload.get("fail_until_attempt") and attempts <= int(payload["fail_until_attempt"]):
            raise RuntimeError(f"{canonical_kind} transient failure on attempt {attempts}")
        if canonical_kind == "write_ledger":
            return self.write_ledger(payload, attempts)
        if canonical_kind == "send_email":
            return self.send_email(payload, attempts)
        if canonical_kind == "create_alert":
            return self.create_alert(payload, attempts)
        if canonical_kind == "create_handoff_summary":
            return self.create_handoff_summary(payload, attempts)
        if canonical_kind == "lookup_resource":
            return self.append_jsonl("resource-lookups.jsonl", canonical_kind, payload, attempts, "resource-lookup")
        if canonical_kind == "follow_up_suggestion":
            return self.append_jsonl("follow-up-suggestions.jsonl", canonical_kind, payload, attempts, "admin-follow-up-review")
        return self.append_jsonl("generic-tool-results.jsonl", canonical_kind, payload, attempts, "generic-local-tool")

    def write_ledger(self, payload: dict[str, Any], attempts: int) -> dict[str, Any]:
        path = self.settings.resolve_path(self.settings.excel_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        report_id = str(payload.get("report_id", ""))
        case_id = str(payload.get("case_id", ""))
        with EXCEL_WRITE_LOCK:
            if path.exists():
                workbook = load_workbook(path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Aegis Risk Ledger"
                sheet.append(["createdAt", "reportId", "caseId", "sessionId", "riskLevel", "summary", "attempts"])
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[1] or "") == report_id and str(row[2] or "") == case_id:
                    return {"delivered_to": "excel-ledger", "path": str(path), "attempts": attempts, "idempotent": True}
            sheet.append(
                [
                    _now_iso(),
                    report_id,
                    case_id,
                    str(payload.get("session_id", "")),
                    str(payload.get("risk_level", "")),
                    str(payload.get("summary", "")),
                    attempts,
                ]
            )
            workbook.save(path)
        self.append_jsonl("ledger-audit.jsonl", "write_ledger", payload, attempts, "excel-ledger")
        return {"delivered_to": "excel-ledger", "path": str(path), "attempts": attempts, "idempotent": False}

    def create_alert(self, payload: dict[str, Any], attempts: int) -> dict[str, Any]:
        result = self.append_jsonl("alert-records.jsonl", "create_alert", payload, attempts, "alert-record")
        webhook_url = self.settings.alert_webhook_url.strip()
        if webhook_url:
            body = json.dumps({"type": "aegis_risk_alert", "payload": redacted_payload(payload)}, ensure_ascii=False).encode("utf-8")
            req = request.Request(webhook_url, data=body, headers={"Content-Type": "application/json"}, method="POST")
            with request.urlopen(req, timeout=self.settings.smtp_timeout_seconds) as response:
                result["webhook_status"] = response.status
        return result

    def send_email(self, payload: dict[str, Any], attempts: int) -> dict[str, Any]:
        mode = self.settings.alert_email_delivery_mode.strip().lower()
        if mode == "log":
            return self.append_jsonl("email-outbox.jsonl", "send_email", payload, attempts, "email-log-outbox")
        if mode != "smtp":
            raise RuntimeError(f"unsupported alert_email_delivery_mode={self.settings.alert_email_delivery_mode}")
        missing = self._missing_smtp_config()
        if missing:
            raise RuntimeError(f"missing SMTP config: {', '.join(missing)}")
        message = self._email_message(payload)
        context = ssl.create_default_context()
        if self.settings.smtp_use_ssl:
            with smtplib.SMTP_SSL(
                self.settings.smtp_host,
                self.settings.smtp_port,
                timeout=self.settings.smtp_timeout_seconds,
                context=context,
            ) as server:
                self._send_message(server, message)
        else:
            with smtplib.SMTP(self.settings.smtp_host, self.settings.smtp_port, timeout=self.settings.smtp_timeout_seconds) as server:
                server.ehlo()
                if self.settings.smtp_use_tls:
                    server.starttls(context=context)
                    server.ehlo()
                self._send_message(server, message)
        self.append_jsonl("email-outbox.jsonl", "send_email", payload | {"delivery_mode": "smtp"}, attempts, "smtp")
        return {"delivered_to": "smtp", "recipient": self.settings.alert_email_to, "attempts": attempts}

    def create_handoff_summary(self, payload: dict[str, Any], attempts: int) -> dict[str, Any]:
        output_dir = self.output_dir / "handoff"
        output_dir.mkdir(parents=True, exist_ok=True)
        case_id = str(payload.get("case_id") or "case")
        path = output_dir / f"{_safe_id(case_id)}.md"
        text = "\n".join(
            [
                "# Aegis Counselor Handoff",
                "",
                f"- Report: {payload.get('report_id', '')}",
                f"- Case: {payload.get('case_id', '')}",
                f"- Risk: {payload.get('risk_level', '')}",
                "",
                "## Summary",
                str(payload.get("summary") or payload.get("suggestion") or ""),
                "",
                "## Next Step",
                "请管理员或辅导员复核学生当前安全状态、可联系支持者和线下交接安排。",
            ]
        )
        path.write_text(text, encoding="utf-8")
        self.append_jsonl("handoff-summaries.jsonl", "create_handoff_summary", payload | {"handoff_path": str(path)}, attempts, "handoff-summary")
        return {"delivered_to": "handoff-summary", "path": str(path), "attempts": attempts}

    def append_jsonl(self, filename: str, kind: str, payload: dict[str, Any], attempts: int, delivered_to: str) -> dict[str, Any]:
        path = self.output_dir / filename
        path.parent.mkdir(parents=True, exist_ok=True)
        row = {
            "created_at": _now_iso(),
            "kind": kind,
            "attempts": attempts,
            "payload": redacted_payload(payload),
        }
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
        return {"delivered_to": delivered_to, "path": str(path), "attempts": attempts}

    @property
    def output_dir(self) -> Path:
        path = self.settings.resolve_path(self.settings.tool_output_dir)
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _missing_smtp_config(self) -> list[str]:
        missing = []
        if not self.settings.smtp_host.strip():
            missing.append("SMTP_HOST")
        if not self._sender():
            missing.append("ALERT_EMAIL_FROM or SMTP_USERNAME")
        if not self._recipients():
            missing.append("ALERT_EMAIL_TO")
        return missing

    def _email_message(self, payload: dict[str, Any]) -> EmailMessage:
        message = EmailMessage()
        message["Subject"] = f"{self.settings.alert_email_subject_prefix} reportId={payload.get('report_id', '')}"
        message["From"] = self._sender()
        message["To"] = ", ".join(self._recipients())
        message.set_content(
            "\n".join(
                [
                    "Aegis 检测到一条高风险心理预警，请尽快安排辅导员或管理员跟进。",
                    "",
                    f"报告ID：{payload.get('report_id', '')}",
                    f"个案ID：{payload.get('case_id', '')}",
                    f"风险等级：{payload.get('risk_level', '')}",
                    f"摘要：{payload.get('summary', '')}",
                    "",
                    "该邮件来自 Aegis 工具执行服务。学生端不会看到后台风险元数据。",
                ]
            )
        )
        return message

    def _send_message(self, server: smtplib.SMTP, message: EmailMessage) -> None:
        if self.settings.smtp_username:
            server.login(self.settings.smtp_username, self.settings.smtp_password)
        server.send_message(message)

    def _sender(self) -> str:
        return self.settings.alert_email_from.strip() or self.settings.smtp_username.strip()

    def _recipients(self) -> list[str]:
        return [item.strip() for item in self.settings.alert_email_to.replace(";", ",").split(",") if item.strip()]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_id(value: str) -> str:
    return "".join(ch for ch in value if ch.isalnum() or ch in {"-", "_"})[:80] or "handoff"
