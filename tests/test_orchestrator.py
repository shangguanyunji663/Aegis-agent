from pathlib import Path
import time

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.config import Settings
from app.database import Base
from app.llm import MockLLMClient
from app.models import ReportStatus, RiskLevel, RuntimeEventType
from app.orchestrator import PsychOrchestrator
from app.repository import DatabaseStore
from app.services.tool_queue import ToolQueueWorker
from app.skills import SkillRegistry


def build_orchestrator(tmp_path: Path) -> PsychOrchestrator:
    knowledge_dir = tmp_path / "knowledge"
    knowledge_dir.mkdir()
    (knowledge_dir / "crisis.md").write_text("自杀 轻生 危机干预 联系可信任的人", encoding="utf-8")
    (knowledge_dir / "exam.md").write_text("考试压力 睡不着 焦虑 可以先稳定身体反应并拆分任务", encoding="utf-8")
    engine = create_engine(f"sqlite:///{tmp_path / 'test.sqlite'}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'test.sqlite'}",
        tool_output_dir=str(tmp_path / "tool-outputs"),
        excel_path=str(tmp_path / "tool-outputs" / "aegis-risk-ledger.xlsx"),
        alert_email_delivery_mode="log",
    )
    store = DatabaseStore(session_factory, settings=settings)
    store.seed_knowledge_dir(knowledge_dir)
    registry = SkillRegistry(knowledge_dir, store.add_report, store.search_knowledge)
    return PsychOrchestrator(registry, store)


def test_companion_intent_skips_knowledge_retrieval(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("今天只是想找人说说话")

    assert response.intent.value == "companion"
    assert all(skill.name != "search_knowledge" for skill in response.skills)
    assert any(item.action == "skip_knowledge" for item in response.trace)


def test_counseling_message_uses_knowledge(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("我最近考试压力很大，晚上睡不着")
    runtime_types = [event.type for event in orchestrator.last_runtime_events]

    assert response.intent.value == "counseling"
    assert response.risk_level is RiskLevel.LOW
    assert any(skill.name == "search_knowledge" for skill in response.skills)
    assert response.pending_report is None
    assert RuntimeEventType.RUN_STARTED in runtime_types
    assert RuntimeEventType.RISK_ASSESSED in runtime_types
    assert RuntimeEventType.ROUTE_DECIDED in runtime_types
    assert RuntimeEventType.KNOWLEDGE_RETRIEVED in runtime_types
    assert RuntimeEventType.SKILLS_SELECTED in runtime_types
    assert RuntimeEventType.RUN_COMPLETED in runtime_types
    assert any(item.agent == "SkillRegistry" and "sleep_routine_support" in item.detail for item in response.trace)
    assert response.response_plan is not None
    assert response.response_plan.response_agent in {"CounselorAgent", "CompanionAgent"}
    assert any(item.action == "compose_plan" for item in response.trace)


def test_high_risk_message_creates_pending_report(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("我不想活了，想结束生命")

    assert response.intent.value == "risk"
    assert response.risk_level is RiskLevel.HIGH
    assert response.pending_report is not None
    assert response.pending_report.status.value == "pending"
    assert any(skill.side_effect for skill in response.skills)


def test_messages_reports_and_traces_are_persisted(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("我不想活了，想结束生命")
    session = orchestrator.store.get_session(response.session_id)
    reports = orchestrator.store.list_reports()
    traces = orchestrator.store.list_traces()

    assert session is not None
    assert [message["role"] for message in session["messages"]] == ["USER", "ASSISTANT"]
    assert len(reports) == 1
    assert reports[0]["status"] == "pending"
    assert len(traces) == 1
    assert traces[0]["intent"] == "risk"


def test_approved_high_risk_report_creates_case_and_notes(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("我不想活了，想结束生命")
    report = orchestrator.store.list_reports()[0]
    updated = orchestrator.store.update_report(report["id"], ReportStatus.APPROVED)
    cases = orchestrator.store.list_cases()
    noted = orchestrator.store.add_case_note(cases[0]["id"], "已联系辅导员跟进", "admin")

    assert updated["status"] == "approved"
    assert updated["emotion"] == "high_risk"
    assert updated["confidence"] >= 0.9
    assert len(cases) == 1
    assert cases[0]["report_id"] == response.pending_report.id
    assert noted["notes"][0]["note"] == "已联系辅导员跟进"


def test_approved_report_creates_runnable_tool_jobs(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    orchestrator.handle("我不想活了，想结束生命")
    report = orchestrator.store.list_reports()[0]
    orchestrator.store.update_report(report["id"], ReportStatus.APPROVED)
    pending = orchestrator.store.list_tool_jobs()
    result = orchestrator.store.run_pending_tool_jobs()
    completed = orchestrator.store.list_tool_jobs()

    assert len(pending) == 5
    assert {job["status"] for job in pending} == {"pending"}
    assert len(result["processed"]) == 5
    assert {job["status"] for job in completed} == {"success"}
    assert any(job["kind"] == "follow_up_suggestion" for job in pending)
    ledger_jobs = [job for job in completed if job["kind"] == "write_ledger"]
    assert ledger_jobs
    ledger_path = Path(ledger_jobs[0]["payload"]["result"]["path"])
    assert ledger_path.suffix == ".xlsx"
    assert ledger_path.exists()
    workbook = load_workbook(ledger_path)
    assert workbook.active.max_row >= 2
    assert any(Path(job["payload"]["result"]["path"]).exists() for job in completed if job["kind"] == "send_email")
    excel_records = orchestrator.store.list_excel_records()
    alert_records = orchestrator.store.list_alert_records()
    assert len(excel_records) == 1
    assert excel_records[0]["status"] == "success"
    assert any(record["channel"] == "alert" and record["status"] == "success" for record in alert_records)
    assert any(record["channel"] == "email" and record["status"] == "success" for record in alert_records)


def test_tool_jobs_can_retry_and_dead_letter(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)
    orchestrator.store.settings.tool_queue_retry_delay_seconds = 0

    job = orchestrator.store.create_tool_job("send_email", {"always_fail": True}, max_attempts=2)
    first = orchestrator.store.run_pending_tool_jobs()
    second = orchestrator.store.run_pending_tool_jobs()

    assert job["status"] == "pending"
    assert any(item["status"] == "pending" for item in first["jobs"] if item["id"] == job["id"])
    assert any(item["status"] == "dead" for item in second["jobs"] if item["id"] == job["id"])
    assert any(item["id"] == job["id"] and item["dead_letter"] for item in orchestrator.store.list_dead_letters())
    retried = orchestrator.store.retry_tool_job(job["id"])
    assert retried is not None
    assert retried["status"] == "pending"


def test_send_email_waits_for_case_tool_dependencies(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)
    orchestrator.store.settings.tool_queue_retry_delay_seconds = 0

    payload = {"report_id": "rep-1", "case_id": "case-1", "risk_level": "high", "summary": "dependency check"}
    email = orchestrator.store.create_tool_job("send_email", payload)
    first = orchestrator.store.run_pending_tool_jobs()
    ledger = orchestrator.store.create_tool_job("write_ledger", payload)
    alert = orchestrator.store.create_tool_job("create_alert", payload)

    second = orchestrator.store.run_pending_tool_jobs()
    jobs = {job["id"]: job for job in second["jobs"]}

    assert email["id"] in first["processed"]
    assert jobs[ledger["id"]]["status"] == "success"
    assert jobs[alert["id"]]["status"] == "success"
    assert jobs[email["id"]]["status"] == "success"
    assert any(item["job_id"] == email["id"] and item["decision"] == "deferred" for item in orchestrator.store.list_tool_audits())


def test_background_tool_worker_processes_pending_jobs_without_manual_run(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)
    orchestrator.store.settings.tool_queue_poll_interval_seconds = 0.05
    orchestrator.store.settings.tool_queue_retry_delay_seconds = 0.05
    orchestrator.store.settings.tool_queue_worker_threads = 2

    orchestrator.handle("我不想活了，想结束生命")
    report = orchestrator.store.list_reports()[0]
    orchestrator.store.update_report(report["id"], ReportStatus.APPROVED)
    worker = ToolQueueWorker(orchestrator.store.settings, orchestrator.store.db_factory)
    try:
        worker.start()
        deadline = time.time() + 5
        while time.time() < deadline:
            jobs = orchestrator.store.list_tool_jobs()
            related = [job for job in jobs if job["report_id"] == report["id"]]
            if related and all(job["status"] == "success" for job in related):
                break
            time.sleep(0.05)
        else:
            assert False, f"background worker did not complete jobs: {orchestrator.store.list_tool_jobs()}"
    finally:
        worker.stop()

    assert orchestrator.store.list_excel_records()
    assert any(record["channel"] == "email" for record in orchestrator.store.list_alert_records())


def test_tool_governance_rejects_unapproved_or_unknown_jobs(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    try:
        orchestrator.store.create_tool_job("send_email", {"risk_level": "high"}, approved=False)
        assert False, "unapproved email job should be rejected"
    except ValueError as exc:
        assert "requires admin approval" in str(exc)

    try:
        orchestrator.store.create_tool_job("unknown_tool", {"risk_level": "high"})
        assert False, "unknown tool should be rejected"
    except ValueError as exc:
        assert "unknown governed tool" in str(exc)

    audits = orchestrator.store.list_tool_audits()
    assert any(item["decision"] == "rejected" and item["tool_kind"] == "send_email" for item in audits)
    assert any(item["decision"] == "rejected" and item["tool_kind"] == "unknown_tool" for item in audits)


def test_tool_payload_includes_contract_and_redaction(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    job = orchestrator.store.create_tool_job(
        "lookup_resource",
        {"risk_level": "low", "message": "我想找学校心理中心", "student_name": "张三"},
        approved=True,
    )

    assert job["payload"]["tool_contract"] == "lookup_resource"
    assert sorted(job["payload"]["redacted_fields"]) == ["message", "student_name"]
    assert job["payload"]["redacted_payload"]["message"] == "[redacted]"
    assert any(item["job_id"] == job["id"] and item["decision"] == "allowed" for item in orchestrator.store.list_tool_audits())


def test_audit_payload_is_redacted(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    orchestrator.store.add_audit_log(
        "usr-test",
        "admin",
        "admin",
        "manual_note",
        "case",
        "case-test",
        {"message": "学生原始求助内容", "phone": "123456", "status": "reviewed"},
    )
    audit = orchestrator.store.list_audit_logs()[0]

    assert audit["payload"]["message"] == "[redacted]"
    assert audit["payload"]["phone"] == "[redacted]"
    assert audit["payload"]["status"] == "reviewed"


def test_knowledge_is_seeded_and_searchable(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    status = orchestrator.store.knowledge_status()
    results = orchestrator.store.search_knowledge("考试压力睡不着")
    chunks = orchestrator.store.ingest_knowledge("custom.md", "人际关系冲突时，可以先描述事实、感受和请求。")

    assert status["database_chunks"] >= 2
    assert "exam.md" in status["sources"]
    assert results[0]["source"] == "exam.md"
    assert chunks == 1
    assert "custom.md" in orchestrator.store.knowledge_status()["sources"]


def test_knowledge_metadata_is_preserved_and_filterable(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    chunks = orchestrator.store.ingest_knowledge(
        "panic.md",
        """---
topic: panic
audience: student
risk_level: medium
source_type: local_markdown
last_reviewed: 2026-07-09
---

惊恐发作时可以先进行 grounding，并联系可信任的人。""",
    )
    status = orchestrator.store.knowledge_status()
    matched = orchestrator.store.search_knowledge("惊恐 grounding", topic="panic", risk_level="medium", audience="student")
    filtered_out = orchestrator.store.search_knowledge("惊恐 grounding", topic="sleep")

    assert chunks == 1
    assert "panic" in status["metadata"]["topics"]
    assert matched[0]["source"] == "panic.md"
    assert matched[0]["metadata"]["topic"] == "panic"
    assert filtered_out == []


def test_mock_llm_boundary_uses_fallback_without_network(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("我最近考试压力很大，晚上睡不着")

    assert isinstance(orchestrator.llm_client, MockLLMClient)
    assert "我听到了你的困扰" in response.answer
    assert any(item.action == "compose_answer" and item.detail == "fallback:mock" for item in response.trace)


def test_memory_summary_is_persisted_and_reused(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    first = orchestrator.handle("我最近考试压力很大，晚上睡不着")
    second = orchestrator.handle("刚才那个问题还是让我很紧张", first.session_id)
    session = orchestrator.store.get_session(first.session_id)

    assert first.memory_summary
    assert second.memory_used is True
    assert "考试压力" in second.answer
    assert session["memory_summary"]
    assert any(item.action == "load_memory" for item in second.trace)


def test_autonomous_agents_persist_private_memory_and_model_profiles(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    response = orchestrator.handle("我最近考试压力很大，晚上睡不着")
    memories = orchestrator.store.load_agent_private_memory("KnowledgeAgent", response.session_id)
    profiles = orchestrator.store.list_agent_model_profiles()

    assert memories
    assert any("context intent" in item["content"] for item in memories)
    assert {item["agent_name"] for item in profiles} >= {"MemoryAgent", "RiskGuardianAgent", "KnowledgeAgent", "CounselorAgent", "CompanionAgent"}
    assert orchestrator.model_registry.profile_for("CounselorAgent")["provider"] == "inherit"


def test_standard_skill_docs_are_loaded(tmp_path: Path):
    orchestrator = build_orchestrator(tmp_path)

    names = orchestrator.registry.standard_skill_names()
    status = orchestrator.registry.standard_skill_status()

    assert "supportive_response_baseline" in names
    assert any(item["name"] == "high_risk_safety_plan" and item["status"] == "ready" for item in status)
